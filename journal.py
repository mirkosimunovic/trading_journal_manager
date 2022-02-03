import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog
import tkinter.messagebox as msgbox
from idlelib.tooltip import Hovertip
#import xlsxwriter
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import glob
import os  
import time
columns = ['Time','Symbol','Order','Risk','Category','MindState','Entry Analysis','Management']


#############  User Defined Options  #######################################################

sheetfilename = 'trading_journal'
categories = ['Med Prob','High Prob','Low Prob']   # customize the levels. Edit, add or remove.
mindstate = ['Normal','Good','Bad']                # customize the levels. Edit, add or remove.

imgdir = os.path.join(os.path.expanduser("~"),'Desktop')	# Can be changed to something like '/Users/my_user_name/Desktop/'   
row_height_pts = 200      # Height of cell rows. This affects the scaling of the chart image
width_info_column = 40    # Width of columns with extended text
#############################################################################################


os.system('rm trading_journal.xlsx')

def image_found(path):
	list_files = glob.glob(path)
	if len(list_files)==0:
		return False
	return True


class Panel(tk.Tk):
	def __init__(self):
		super().__init__()
		self.title("Trading Journal Manager v1.0")
		self.geometry("+50+50")
		self.resizable(0,0)
		style = ttk.Style(self)
		style.configure('TNotebook.Tab', padding=(7, 6, 7, 0))		
		estyle = ttk.Style()
		estyle.configure("EntryStyle.TEntry", background='white')		
		#style.theme_use('classic')   # aqua, alt, default, clam, classic

		# variables
		self.sheetfilename = sheetfilename+'.xlsx'
		self.logging_dict = {key:tk.StringVar() for key in columns}
		self.spreadsheet = Spreadsheet(self.sheetfilename)
		self.opentrades,self.opentrade_rownum = self.get_open_trades()
		self.open_trade_id = tk.StringVar()
		self.num_open_trades = tk.StringVar()
		self.num_open_trades.set('You have {} open trades'.format(sum([el!='None' for el in self.opentrades])))
		global screenshot_format; screenshot_format = tk.StringVar()
		screenshot_format.set('*.png')
		global use_screenshot; use_screenshot = tk.IntVar()
		use_screenshot.set(1)
		global new_screenshot; new_screenshot = tk.IntVar()
		new_screenshot.set(0)
		global imgpath; imgpath = os.path.join(imgdir,screenshot_format.get())


		# Create the Header Label
		self.label = ttk.Label(self,textvariable=self.num_open_trades,font=('Serif Bold',20),foreground='black',anchor=tk.CENTER)
		self.label.grid(sticky='news',columnspan=5,pady=5)

		n = ttk.Notebook(self)
		tab_1 = ttk.Frame(n)
		tab_2 = ttk.Frame(n)
		tab_3 = ttk.Frame(n)
		n.add(tab_1,text=" New Trade")
		n.add(tab_2,text=" Open Trades")
		n.add(tab_3,text=" Options")
		n.grid(sticky=tk.NSEW,columnspan=5)
		#n.pack(expand=1, fill ="both")
		#tab_1.columnconfigure( (0,1), weight=1)
		#tab_1.rowconfigure((0,1), weight=1)

		##############################################################################################
		# TAB 1  New Trades #########################################################################

		ttk.Label(tab_1,text="Symbol:").grid(row=0,column=0,sticky='news',padx=(5,0),pady=10)	
		ttk.Entry(tab_1,textvariable=self.logging_dict['Symbol'],style='EntryStyle.TEntry',width=7).grid(row=0,column=1,sticky='news',padx=(1,5),pady=10)
		order = ttk.OptionMenu(tab_1, self.logging_dict['Order'],"Buy","Buy","Sell")
		order.grid(row=0,column=2,padx=3,pady=10,sticky='news')
		order.config(width=4)
		ttk.Label(tab_1,text="Risk:").grid(row=0,column=3,sticky='news',padx=(3,0),pady=10)
		risk_entry = ttk.Entry(tab_1,textvariable=self.logging_dict['Risk'],style='EntryStyle.TEntry',width=2)
		risk_entry.grid(row=0,column=4,sticky='news',padx=0,pady=10)		
		risk_entry.insert(0,"1")
		ttk.Label(tab_1,text="%").grid(row=0,column=5,sticky='news',padx=0,pady=10)

		ttk.Label(tab_1,text="Category:").grid(row=1,column=0,sticky='news',padx=(5,0),pady=0)
		ttk.Label(tab_1,text="Mind State:").grid(row=1,column=2,sticky='news',padx=(5,0),pady=0)				
		ttk.OptionMenu(tab_1, self.logging_dict['Category'],categories[0],*categories).grid(row=2,column=0,columnspan=2,padx=5,pady=0,sticky='news')
		ttk.OptionMenu(tab_1, self.logging_dict['MindState'],mindstate[0],*mindstate).grid(row=2,column=2,columnspan=2,padx=5,pady=0,sticky='news')
		
		ttk.Label(tab_1,text="Entry Analysis:").grid(row=3,column=0,columnspan=2,sticky='news',padx=5,pady=1)
		self.analysis = tk.Text(tab_1,highlightcolor="LightSteelBlue2",width = width_info_column, height = 10)
		self.analysis.grid(row=4,column=0,columnspan=5,sticky='news',padx=5,pady=1)
		
		ttk.Label(tab_1,text="Management Rules:").grid(row=5,column=0,columnspan=2,sticky='news',padx=5,pady=1)
		self.mgmt = tk.Text(tab_1,highlightcolor="LightSteelBlue2",width = width_info_column, height = 10)
		self.mgmt.grid(row=6,column=0,columnspan=5,sticky='news',padx=5,pady=1)		
		
		style.configure('my.TButton',foreground = 'blue')		
		ttk.Button(tab_1, style='my.TButton', text='Add Entry', command=self.add_entry).grid(padx=5,pady=1,sticky='news',column=1,columnspan=2)
		ttk.Checkbutton(tab_1,text="attach last screenshot",variable=use_screenshot).grid(padx=5,pady=1,sticky='news',column=0,columnspan=2)

		##############################################################################################
		# TAB 2  Open Trades #########################################################################

		ttk.Label(tab_2,text="Select:").grid(row=0,column=0,sticky='news',padx=(5,0),pady=10)	
		self.openMenu = ttk.OptionMenu(tab_2, self.open_trade_id,self.opentrades[0],*self.opentrades)
		self.openMenu.grid(row=0,column=1,padx=3,pady=10,sticky='news',columnspan=3)		
		self.open_trade_id.trace("w",self.OptionMenu_SelectionEvent)

		ttk.Label(tab_2,text="Result/Comments:").grid(row=1,column=0,columnspan=2,sticky='news',padx=5,pady=5)
		self.result = tk.Text(tab_2,highlightcolor="LightSteelBlue2",width = width_info_column, height = 10)
		self.result.grid(row=2,column=0,columnspan=5,sticky='news',padx=5,pady=1)

		ttk.Button(tab_2, style='my.TButton', text='Update Open Trade', command=self.add_resulttext).grid(padx=5,pady=1,sticky='news',row=3,column=0,columnspan=2)
		ttk.Checkbutton(tab_2,text="update with last screenshot",variable=new_screenshot).grid(padx=5,pady=1,sticky='news',row=4,column=0,columnspan=2)

		s2 = ttk.Style()
		s2.configure('my2.TButton',foreground = 'blue',font=('Sans',12,'bold'))		
		ttk.Button(tab_2, style='my2.TButton', text='Close Trade', command=self.close_opentrade).grid(padx=5,pady=20,sticky='news',row=5,column=2,columnspan=2)


		##############################################################################################
		# TAB 3 OPTIONS #################################################################################

		ttk.Label(tab_3,text="Screenshot format:").grid(row=0,column=0,sticky='news',padx=(30,5),pady=(20,1))	
		ttk.OptionMenu(tab_3, screenshot_format,screenshot_format.get(),"*.png","*.jpg","*.jpeg").grid(row=0,column=1,sticky='news',padx=5,pady=(20,1))

		# Create the 'Load Folder' button.
		loaddir = ttk.Button(tab_3, text='Load Screenshots Folder', command=self.load_dir)
		loaddir.grid(row=1,column=0,sticky='news', padx=30,pady=5,columnspan=2)
		Hovertip(loaddir, "Load the directory to locate screenshots",hover_delay=500)


		##############################################################################################
		# Functions   ################################################################################

	def add_entry(self):

		if use_screenshot.get() and not image_found(imgpath):
			tk.messagebox.showerror("Error","Cannot find screenshot files.\nGo to Options tab and select correct settings.")
			return

		now = time.strftime('%a, %b-%d %H:%M')
		self.logging_dict['Entry Analysis'].set(self.analysis.get("1.0",'end-1c')) # get text from line 1 character 0, until end minus 1 character (removes extra line)
		self.logging_dict['Management'].set(self.mgmt.get("1.0",'end-1c'))    
		self.logging_dict['Time'].set(now) 
		if any([self.logging_dict[key].get()=="" for key in self.logging_dict]):
			tk.messagebox.showerror("Error", "Please fill all fields.")
			return 
		self.spreadsheet.add_entry(self.logging_dict)
		self.analysis.delete("1.0","end")
		self.mgmt.delete("1.0","end")

		self.update_panel()

	def get_open_trades(self):
		sh = self.spreadsheet.sheet
		maxrow = self.spreadsheet.numrows
		opentrades = []
		row_number = []
		if maxrow<2:
			return ['None',],[0,]
		else:
			for row in range(2,maxrow+1):
				time,symbol,order,closed = sh.cell(row=row,column=1).value,sh.cell(row=row,column=2).value,\
				                           sh.cell(row=row,column=3).value,sh.cell(row=row,column=11).value
				this_trade = time+' '+symbol+' '+order
				if closed=='X':
					continue
				row_number.append(row)
				opentrades.append(this_trade)

		return (opentrades,row_number) if opentrades!=[] else (['None',],[0,])

	def add_resulttext(self):

		if new_screenshot.get() and not image_found(imgpath):
			tk.messagebox.showerror("Error","Cannot find screenshot files.\nGo to Options tab and select correct settings.")
			return			
		if self.open_trade_id.get()=='None':
			return
		text = self.result.get("1.0",'end-1c')
		self.spreadsheet.add_text(text,self.opencellrow,10,align_horiz='left')

		if new_screenshot.get():
			self.add_newchart(self.opencellrow)

	def add_newchart(self,row):
		
		self.spreadsheet.add_chart(row)			

	def close_opentrade(self):

		if self.open_trade_id.get()=='None':
			return
		self.spreadsheet.add_text('X',self.opencellrow,11,align_horiz='center')
		self.update_panel()		

	def load_dir(self):

		imgdir = tkinter.filedialog.askdirectory(title='Select screenshots folder',initialdir='./')
		global imgpath; imgpath = os.path.join(imgdir,screenshot_format.get())	

	def OptionMenu_SelectionEvent(self,*args):

		index_ = self.opentrades.index(self.open_trade_id.get())
		self.opencellrow = self.opentrade_rownum[index_]	
		self.result.delete("1.0","end")
		self.result.insert("1.0",self.spreadsheet.get_text(self.opencellrow,10))

	def update_panel(self):

		self.opentrades,self.opentrade_rownum = self.get_open_trades()	
		self.openMenu.set_menu(self.opentrades[0],*self.opentrades)
		self.num_open_trades.set('You have {} open trades'.format(sum([el!='None' for el in self.opentrades])))	
		self.update()



class Spreadsheet():

	def __init__(self,sheetfilename):
		self.filepath = os.getcwd()		# spreadhseet are created in current working directory
		self.fullpath = os.path.join(self.filepath,sheetfilename)
		self.workbook = self.get_workbook()
		self.update()

	def get_workbook(self):

		file_exists = os.path.exists(self.fullpath)
		if file_exists:
			return openpyxl.load_workbook(self.fullpath)
		else:
			newbook = openpyxl.Workbook()
			newbook = self.make_sheet_header(newbook)
			return newbook

	def make_sheet_header(self,book):
		sheet  = book.active
		columns_ = columns+['Chart','Result/Comments','Closed']

		# Fill the header
		for ix,col in enumerate(columns_):
			sheet.cell(row=1,column=ix+1).value = col
			sheet.cell(row=1,column=ix+1).font = Font(bold=True)
			sheet.cell(row=1,column=ix+1).alignment = Alignment(vertical="center",horizontal="center")
			sheet.column_dimensions[get_column_letter(ix+1)].width = len(col)*1.4

		# Set the column widths
		for col in ['Entry Analysis','Management','Result/Comments']:
			indx = columns_.index(col)+1
			sheet.column_dimensions[get_column_letter(indx)].width = width_info_column
		sheet.column_dimensions[get_column_letter(columns_.index('Time')+1)].width = 16	
		sheet.column_dimensions[get_column_letter(columns_.index('Chart')+1)].width = width_info_column+10	

		# Set the header background color
		for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
			for cell in rows:
				cell.fill = PatternFill(start_color='FF27E85B',end_color='FF27E85B', fill_type = "solid")

		# Set the header row height
		rd = sheet.row_dimensions[1]
		rd.height = 20		

		book.save(self.fullpath)
		return openpyxl.load_workbook(self.fullpath)

	def add_entry(self,logging_dict):

		for ix,key in enumerate(logging_dict):
			self.sheet.cell(row=self.numrows+1,column=ix+1).value = logging_dict[key].get()
			self.sheet.cell(row=self.numrows+1,column=ix+1).alignment = Alignment(vertical="top",wrapText=True)

		# Set the cell row height for log cells		
		rd = self.sheet.row_dimensions[self.numrows+1]
		rd.height = row_height_pts

		if use_screenshot.get():	# case when attaching screenshot
			self.add_chart(self.numrows+1)
		else:						# case when not using screenshot
			self.update()
			return

	def add_chart(self,row):

		# Find screenshot image
		list_of_images = glob.glob(imgpath) 
		latest_image = max(list_of_images, key=os.path.getctime)
		# Open image and anchor to worksheet
		img = openpyxl.drawing.image.Image(latest_image)
		img.anchor = self.sheet.cell(row=row,column=9).coordinate
		img.width = img.width*(row_height_pts/img.height)*1.333*0.95
		img.height = row_height_pts*1.333*0.95
		self.sheet.add_image(img)		
		self.update()

	def add_text(self,text,row,col,align_horiz="left"):

		self.sheet.cell(row=row,column=col).value = text
		self.sheet.cell(row=row,column=col).alignment = Alignment(vertical="top",horizontal=align_horiz,wrapText=True)
		self.update()

	def get_text(self,row,col):

		if row<1:
			return ""

		text = self.sheet.cell(row=row,column=col).value
		return text if text is not None else ""

	def update(self):

		self.workbook.save(self.fullpath)
		self.workbook = openpyxl.load_workbook(self.fullpath)		
		self.sheet = self.workbook.active
		self.numrows = self.sheet.max_row





if __name__ == '__main__':
	Panel = Panel()
	Panel.mainloop()